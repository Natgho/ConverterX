# Created by SezerBozkir<admin@sezerbozkir.com> at 2/19/2020
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter.ttk import Progressbar
from openpyxl import Workbook
import csv


def get_csv():
    global file_path

    file_path = filedialog.askopenfilename()


def convert_to_excel():
    progress['value'] = 0
    root.update_idletasks()
    global file_path, split_size_entry
    split_size = int(split_size_entry.get())

    export_file_path = filedialog.asksaveasfilename()

    rows = []
    with open(file_path, 'r') as f:
        for row in csv.reader(f):
            rows.append(row)
    header = rows.pop(0)
    if split_size > 1:
        remaining_value = len(rows) % split_size
        for part in range(split_size, len(rows), split_size):
            wb = Workbook()
            ws = wb.active
            ws.append(header)
            for row_order in range(part - split_size, part, 1):
                ws.append(rows[row_order])
            wb.save(export_file_path + f"_{part}.xlsx")
        if remaining_value:
            wb = Workbook()
            ws = wb.active
            ws.append(header)
            for rem_index in range(len(rows) - remaining_value, len(rows), 1):
                ws.append(rows[rem_index])
            wb.save(export_file_path + f"_{rem_index + 1}.xlsx")
    else:
        wb = Workbook()
        ws = wb.active
        for order in range(len(rows)):
            ws.append(rows[order])
            wb.save(export_file_path + ".xlsx")

    # https://www.geeksforgeeks.org/progressbar-widget-in-tkinter-python/
    progress['value'] = 100
    root.update_idletasks()


def exit_app():
    MsgBox = tk.messagebox.askquestion('Exit Application', 'Are you sure you want to exit the application',
                                       icon='warning')
    if MsgBox == 'yes':
        root.destroy()


root = tk.Tk(className='ConverterX')
root.resizable(width=False, height=False)
canvas1 = tk.Canvas(root, width=300, height=300, bg='lightsteelblue2', relief='raised')
canvas1.pack()

label1 = tk.Label(root, text='ConverterX', bg='lightsteelblue2')
label1.config(font=('helvetica', 20))
canvas1.create_window(150, 60, window=label1)
file_path = ""
saveAsButton_Excel = tk.Button(text='Convert CSV to Excel', command=convert_to_excel, bg='green', fg='white',
                               font=('helvetica', 12, 'bold'))
canvas1.create_window(150, 180, window=saveAsButton_Excel)

browseButton_CSV = tk.Button(text="      Import CSV File     ", command=get_csv, bg='green', fg='white',
                             font=('helvetica', 12, 'bold'))
canvas1.create_window(150, 130, window=browseButton_CSV)
split_part_label = tk.Label(text='Split Size:', bg='lightsteelblue2')
canvas1.create_window(90, 215, window=split_part_label)
split_size_entry = tk.Entry(root)
split_size_entry.insert(0, '1')
canvas1.create_window(180, 215, window=split_size_entry)
exitButton = tk.Button(root, text='       Exit Application     ', command=exit_app, bg='brown', fg='white',
                       font=('helvetica', 12, 'bold'))
canvas1.create_window(150, 250, window=exitButton)
progress = Progressbar(root, orient=tk.HORIZONTAL,
                       length=100, mode='determinate')
progress.pack(pady=10)
root.mainloop()
